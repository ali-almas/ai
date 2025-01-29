import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# Get the Downloads directory path dynamically
downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
excel_file = os.path.join(downloads_folder, 'test.xlsx')
new_excel_file = os.path.join(downloads_folder, 'result.xlsx')

# Read the input Excel file
df = pd.read_excel(excel_file)

# Initialize the data dictionary with empty lists
data = {
    'Action': [],
    'Amount': [],
    'Price': [],
    'Total': [],
    'Date': []
}

# Define the handle_buy and handle_sell functions
def handle_buy(row):
    data['Action'].append('Buy')
    # Convert the first part of 'Sell' to a float, no formatting
    data['Amount'].append(row['Buy'].split()[0])
    data['Price'].append(row['Inverse Price'].split("=")[1].split()[0])
    data['Total'].append(row['Sell'])
    # Convert Date format
    date_obj = pd.to_datetime(row['Date']).strftime('%d %B %Y')
    data['Date'].append(date_obj)

def handle_sell(row):
    data['Action'].append('Sell')
    # Convert the first part of 'Sell' to a float, no formatting
    data['Amount'].append(row['Sell'].split()[0])
    data['Price'].append(row['Price'].split("=")[1].split()[0])
    data['Total'].append(row['Buy'])
    # Convert Date format
    date_obj = pd.to_datetime(row['Date']).strftime('%d %B %Y')
    data['Date'].append(date_obj)

# Iterate through the DataFrame
for index, row in df.iterrows():
    if row['Status'] == 'Successful':
        if 'USDT' in row['Sell'] or 'USDC' in row['Sell']:
            handle_buy(row)
        else:
            handle_sell(row)

# Convert the data dictionary into a DataFrame and save to an Excel file in the Downloads folder
output_df = pd.DataFrame(data)

output_df.to_excel(new_excel_file, index=False)

# Open the saved Excel file to manipulate column width
wb = load_workbook(new_excel_file)
ws = wb.active

# Set font size for all cells in the sheet
font = Font(size=16)  # Set the font size to 12

# Apply the font to all cells
for row in ws.iter_rows():
    for cell in row:
        cell.font = font

# Set the column widths (you can adjust these values)
ws.column_dimensions['A'].width = 15  # 'Action' column width
ws.column_dimensions['B'].width = 30  # 'Amount' column width
ws.column_dimensions['C'].width = 30  # 'Price' column width
ws.column_dimensions['D'].width = 30  # 'Total' column width
ws.column_dimensions['E'].width = 30  # 'Date' column width

# Save the changes
wb.save(new_excel_file)

print(f"Excel file saved to '{new_excel_file}' successfully!")
